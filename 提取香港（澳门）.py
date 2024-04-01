import openpyxl

# 打开xlsx文件
workbook = openpyxl.load_workbook("x.xlsx")
sheet = workbook.active

# 创建一个空的列表用于保存含有“香港”字段的单元格内容
hong_kong_cells = []

# 遍历所有单元格
for row in sheet.iter_rows(values_only=True):
    for cell_value in row:
        # 如果单元格内容包含“香港”，则将其保存到列表中
        if isinstance(cell_value, str) and "澳门" in cell_value:
            hong_kong_cells.append(cell_value)

# 创建一个新的xlsx文件来保存矩阵
new_workbook = openpyxl.Workbook()
new_sheet = new_workbook.active

# 将保存的单元格内容写入新的xlsx文件的一列中
for index, value in enumerate(hong_kong_cells, start=1):
    new_sheet.cell(row=index, column=1, value=value)

# 保存新的xlsx文件
new_workbook.save("澳门.xlsx")
