import os
import pandas as pd
import openpyxl
from openpyxl import load_workbook

# 外层文件夹路径
outer_dir = '/mnt/c/大学/科研/大湾区'

# 内层文件夹名称
inner_folders = ['15.3', '21.7','66-(0-12)','66-（12-32）','66-(32-41)','66-(41-54)','66-(54-65)','66-(65-66.8)']

for folder_name in inner_folders:
    # 内层文件夹路径
    inner_dir = os.path.join(outer_dir, folder_name)

    # 获取所有.xlsx文件
    xlsx_files = [file for file in os.listdir(inner_dir) if file.endswith('.xlsx')]

    # 创建一个空的DataFrame来存储所有数据
    all_data = pd.DataFrame()

    # 遍历每个.xlsx文件
    for file_name in xlsx_files:
        # 读取Excel文件
        df = pd.read_excel(os.path.join(inner_dir, file_name))

        selected_data = df[["公开（公告）号", "申请日", "专利类型", "公开类型", "知识密集型分类", "清洁能源产业", "清洁能源产业", "被引证次数", "学科分类", "新兴产业(主)", "申请人"]]

        # 将提取的数据添加到all_data中
        all_data = pd.concat([all_data, selected_data], ignore_index=True)

    # 创建新的Excel文件名
    new_file_name = os.path.join(inner_dir, "所有数据提取.xlsx")

    # 将所有数据保存到新的Excel文件中
    all_data.to_excel(new_file_name, index=False)

    # 打开 Excel 文件
    workbook = openpyxl.load_workbook(new_file_name)
    sheet = workbook.active

    # 遍历申请人列的所有单元格
    for cell in sheet['K']:
        # 检查单元格内容是否包含分号
        if ';' in str(cell.value):
            # 分割单元格内容
            values = str(cell.value).split(';')
            # 将内容分列到右边的单元格中
            for i, value in enumerate(values):
                sheet.cell(row=cell.row, column=cell.column + i).value = value

    # 删除单元格的空格
    for row in sheet.iter_rows():
        for cell in row:
            # 检查单元格的内容是否以空格开头
            if isinstance(cell.value, str) and cell.value.startswith(' '):
                # 去除开头的空格
                cell.value = cell.value.lstrip()

    #删除所有人名
    # 遍历所有单元格，从第 K 列开始检查
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=11, max_col=sheet.max_column):
        for cell in reversed(row):
            # 检查单元格内容长度是否小于等于三
            if len(str(cell.value)) <= 3:
                # 删除单元格的内容
                cell.value = None
                # 将该单元格所在行的右侧所有单元格左移
                for i in range(cell.column + 1, sheet.max_column + 1):
                    sheet.cell(row=cell.row, column=i - 1).value = sheet.cell(row=cell.row, column=i).value
                    sheet.cell(row=cell.row, column=i).value = None

    # 保存修改后的 Excel 文件
    workbook.save(new_file_name)

    # 删除单个申请人行
    df = pd.read_excel(new_file_name)

    # 起始列索引（从0开始）
    start_column_index = 10  # 第K列对应的索引

    # 获取每行非空单元格的个数
    non_empty_counts = df.iloc[:, start_column_index:].notnull().sum(axis=1)

    # 找到非空单元格数量小于2的行索引
    rows_to_delete = non_empty_counts[non_empty_counts < 2].index

    # 删除对应的行
    df = df.drop(rows_to_delete, axis=0)

    # 将修改后的 DataFrame 保存到 Excel 文件中
    df.to_excel(new_file_name, index=False)
