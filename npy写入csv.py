
'''
import openpyxl
import networkx as nx
import itertools
from tqdm import tqdm
import numpy as np

# 读取Excel文件
file_path = "x.xlsx"
wb = openpyxl.load_workbook(file_path)
sheet = wb.active

# 创建一个无向图
G = nx.Graph()

# 统计边的出现次数
edge_counter = {}

# 遍历每一行，构建无向网络
for row in sheet.iter_rows(values_only=True):
    # 过滤空值
    row_values = [value for value in row if value is not None]
    # 添加边
    for i in range(len(row_values)):
        for j in range(i + 1, len(row_values)):
            edge = (row_values[i], row_values[j])
            # 更新边的出现次数
            edge_counter[edge] = edge_counter.get(edge, 0) + 1

# 添加带有权重的边
for edge, weight in edge_counter.items():
    G.add_edge(edge[0], edge[1], weight=weight)

# 写入矩阵
nodes = list(G.nodes())

# 创建带权重的邻接矩阵
num_nodes = len(nodes)
adj_matrix = np.zeros((num_nodes, num_nodes))
for i, node1 in tqdm(enumerate(nodes)):
    for j, node2 in enumerate(nodes):
        if G.has_edge(node1, node2):
            adj_matrix[i, j] = G[node1][node2]['weight']

print(adj_matrix)

# 写入矩阵
nodes = list(G.nodes())
#print(nodes)

# 保存节点名称到CSV文件
with open('node_names.csv', 'w') as f:
    for node in nodes:
        f.write("%s\n" % node)
print("节点名称保存完成为CSV文件")

# 保存带权重的邻接矩阵为CSV文件
np.savetxt('adj_matrix1.csv', adj_matrix, delimiter=',')
print("带权重的邻接矩阵保存完成为CSV文件")
'''

import openpyxl
import networkx as nx
import itertools
from tqdm import tqdm
import numpy as np

# 读取Excel文件
file_path = "x.xlsx"
wb = openpyxl.load_workbook(file_path)
sheet = wb.active

# 创建一个无向图
G = nx.Graph()

# 统计边的出现次数
edge_counter = {}

# 遍历每一行，构建无向网络
for row in sheet.iter_rows(values_only=True):
    # 过滤空值
    row_values = [value for value in row if value is not None]
    # 添加边
    for i in range(len(row_values)):
        for j in range(i + 1, len(row_values)):
            edge = (row_values[i], row_values[j])
            # 更新边的出现次数
            edge_counter[edge] = edge_counter.get(edge, 0) + 1

# 添加带有权重的边
for edge, weight in edge_counter.items():
    G.add_edge(edge[0], edge[1], weight=weight)

# 写入矩阵
nodes = list(G.nodes())

# 创建带权重的邻接矩阵
num_nodes = len(nodes)
adj_matrix = np.zeros((num_nodes, num_nodes))
for i, node1 in tqdm(enumerate(nodes)):
    for j, node2 in enumerate(nodes):
        if G.has_edge(node1, node2):
            adj_matrix[i, j] = G[node1][node2]['weight']

print(adj_matrix)

# 保存节点名称到CSV文件
with open('node_names.csv', 'w', encoding='utf-8-sig') as f:
    for node in nodes:
        f.write("%s\n" % node)
print("节点名称保存完成为CSV文件")

# 保存带权重的邻接矩阵为CSV文件
np.savetxt('adj_matrix1.csv', adj_matrix, delimiter=',')
print("带权重的邻接矩阵保存完成为CSV文件")
