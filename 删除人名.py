from openpyxl import load_workbook

# 加载工作簿
workbook = load_workbook("所有数据提取.xlsx")
# 选择第一个工作表
sheet = workbook.active

# 遍历所有单元格
for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
    for cell in reversed(row):
        # 检查单元格内容长度是否小于等于三
        if len(str(cell.value)) <= 3:
            # 删除单元格的内容
            cell.value = None
            # 将该单元格所在行的右侧所有单元格左移
            for i in range(cell.column + 1, sheet.max_column + 1):
                sheet.cell(row=cell.row, column=i - 1).value = sheet.cell(row=cell.row, column=i).value
                sheet.cell(row=cell.row, column=i).value = None

# 保存修改后的工作簿
workbook.save("所有数据提取_处理后.xlsx")
